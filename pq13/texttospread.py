import openpyxl
import os
from openpyxl.utils import get_column_letter


wb = openpyxl.Workbook()
sheet = wb['Sheet']

list = []
for i in os.listdir('/home/satvshr/Desktop/amfoss-tasks-2/pq13'):
    if i.endswith('.txt'):
        text = open('/home/satvshr/Desktop/amfoss-tasks-2/pq13/' + i)
        list.append(text.readlines())

col = 1
#row = 1
for i in list:
    #print(list)
    row = 1
    for j in i:
        #print(j)
        sheet.cell(row=row, column=col).value = j
        row += 1
    col += 1
wb.save('texttospread.xlsx')
