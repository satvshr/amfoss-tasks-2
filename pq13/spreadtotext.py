import openpyxl

wb = openpyxl.load_workbook('/home/satvshr/Desktop/amfoss-tasks-2/pq13/spreadtotext.xlsx')
sheet = wb['Sheet1']
# Treat each column as a its own new text file.
for c in range(1, sheet.max_column + 1):

    # First row is treated as a header.
    file_name = sheet.cell(row = 1, column = c).value + '.txt'
    
    text_file = open(file_name, 'w')

    lines = []

    # For all subsequent rows after the header.
    for r in range(2, sheet.max_row + 1):

        v = sheet.cell(row = r, column = c).value
    
        # Check if cell has content, then add to list of lines.
        if v is not None:
            lines.append(v)

    text_file.write('\n'.join(lines))
    
    print('Spreadsheet lines from ' + file_name + ' successfully saved ')
    
    text_file.close()