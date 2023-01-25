import openpyxl
open = openpyxl.load_workbook('inverter.xlsx')
sheet = open['Sheet']
sheet1 = open['Sheet1']
rows = sheet.max_row
columns = sheet.max_column
#print(rows)
#print(columns)
for i in range(1, columns+1):
    for j in range(1, rows+1):
        sheet1.cell(row = i, column = j).value = sheet.cell(row = j, column = i).value
open.save('inverter.xlsx')
