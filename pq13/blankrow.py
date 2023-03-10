import openpyxl, sys
if len(sys.argv) == 4:
    n = int(sys.argv[1])
    m = int(sys.argv[2])
    doc = sys.argv[3]

    open = openpyxl.load_workbook(doc)
    sheet = open['Sheet']
    sheet1 = open['Sheet1']
    col_len = 0
    row_len = 0

    for i in range(1, 1000):
        if sheet.cell(row = i, column = 1).value != None:
            row_len += 1
        else:
            break
    for i in range(1, 1000):
        if sheet.cell(row = 1, column = i).value != None:
            col_len += 1
        else:
            break
    #print(sheet.max_row)
    #print(sheet.max_column)
    for i in range(1, n+1):
        for j in range(1, col_len+1):
            sheet1.cell(row = i, column = j).value = sheet.cell(row = i, column = j).value
            # insert all elements till nth row

    for i in range(n+m+1, row_len+m+1): #, n+m+1+row_len-n #n+m+1, row_len+m+1
        for j in range(1, col_len+1):
            sheet1.cell(row = i, column = j).value = sheet.cell(row = i-m, column = j).value
            # insert elements from (n+m+1)th row till (row+m+1)th row

    for i in range(n+1, n+m+1):
        for j in range(1, col_len+1):
            sheet1.cell(row = i, column = j).value = None
    open.save(doc)
else:
    print("Format should be: python3 <your programme name> <rows you want before blank lines> <number of blank lines you want> <table name>")


