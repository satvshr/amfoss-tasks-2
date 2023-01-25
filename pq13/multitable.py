import openpyxl, sys
wb = openpyxl.Workbook()
sheet = wb['Sheet']
if len(sys.argv) == 2:
    n = int(sys.argv[1])

    for i in range(1, n+1):
        sheet.cell(row=i+1, column=1).value = i
    for i in range(1, n+1):
        sheet.cell(row=1, column=i+1).value = i
    for i in range(1, n+1):
        for j in range(1, n+1):
            prod = i*j
            sheet.cell(row=i+1, column=j+1).value = prod
    wb.save('multiplyfoss.xlsx')

else:
    print("Format should be: python3 <your programme name> <number of rows/columns you want>")
